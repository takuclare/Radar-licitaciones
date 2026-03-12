# summarizer.py
import os
import re
import json
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Callable

import requests
from bs4 import BeautifulSoup
from pdfminer.high_level import extract_text
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from dotenv import load_dotenv
from openai import OpenAI
import unicodedata

load_dotenv()

# =========================================================
# OpenAI
# =========================================================
def _get_client() -> OpenAI:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Falta OPENAI_API_KEY en .env")
    return OpenAI(api_key=api_key, timeout=180.0, max_retries=1)


def _clean_inline(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s





def _clean_keep_newlines(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("\r", "\n")
    s = re.sub(r"[\t\f\v]+", " ", s)
    s = re.sub(r" *\n *", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    s = re.sub(r" {2,}", " ", s)
    return s.strip()

def _normalize(s: str) -> str:
    s = (s or "").lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s).strip()
    return s
# =========================================================
# JSON robust parsing
# =========================================================
def _extract_first_json_object(text: str) -> str:
    if not text:
        raise ValueError("Texto vacío")

    start = text.find("{")
    if start == -1:
        raise ValueError("No hay '{' en la respuesta")

    depth = 0
    in_str = False
    esc = False

    for i in range(start, len(text)):
        ch = text[i]

        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue
        else:
            if ch == '"':
                in_str = True
                continue
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return text[start:i + 1]

    raise ValueError("JSON no balanceado (faltan llaves de cierre)")


def _parse_json_or_raise(text: str) -> Dict:
    text = (text or "").strip()
    if not text:
        raise ValueError("Texto vacío")
    try:
        return json.loads(text)
    except Exception:
        obj = _extract_first_json_object(text)
        return json.loads(obj)


def _responses_output_text(resp) -> str:
    """
    Compatibilidad con distintas versiones del SDK: intenta sacar texto de salida.
    """
    txt = (getattr(resp, "output_text", None) or "").strip()
    if txt:
        return txt

    # fallback: recorrer resp.output
    try:
        chunks = []
        for item in getattr(resp, "output", []) or []:
            for c in getattr(item, "content", []) or []:
                t = getattr(c, "text", None)
                if t:
                    chunks.append(t)
        return "\n".join(chunks).strip()
    except Exception:
        return ""


# =========================================================
# Anti-generic / anti-reference helpers
# =========================================================
_BANNED_REFERENCE_PATTERNS = [
    r"conforme\s+al?\s+anexo",
    r"seg[uú]n\s+el?\s+anexo",
    r"de\s+acuerdo\s+con\s+el?\s+anexo",
    r"ver\s+anexo",
    r"v[eé]ase\s+anexo",
    r"consultar\s+anexo",
    r"seg[uú]n\s+pcap",
    r"conforme\s+al?\s+pcap",
    r"de\s+acuerdo\s+con\s+el?\s+pcap",
    r"seg[uú]n\s+pliego",
    r"conforme\s+al?\s+pliego",
    r"de\s+acuerdo\s+con\s+el?\s+pliego",
    r"seg[uú]n\s+el\s+apartado",
    r"conforme\s+al\s+apartado",
    r"ver\s+apartado",
    r"seg[uú]n\s+la\s+cl[aá]usula",
    r"conforme\s+a\s+la\s+cl[aá]usula",
    r"de\s+acuerdo\s+con\s+la\s+cl[aá]usula",
    r"ver\s+cl[aá]usula",
    r"seg[uú]n\s+el\s+cuadro\s+resumen",
    r"ver\s+cuadro\s+resumen",
]

def _has_banned_references(s: str) -> bool:
    t = _clean_inline(s).lower()
    if not t:
        return False
    for p in _BANNED_REFERENCE_PATTERNS:
        if re.search(p, t, flags=re.IGNORECASE):
            return True
    return False

def _fields_need_dereference(data: Dict) -> bool:
    """
    True si el JSON contiene frases referenciales/genericas en campos clave.
    """
    watch_keys = [
        "objeto_detallado",
        "solvencia_tecnica_detallada",
        "medios_humanos_materiales_detallados",
        "criterios_adjudicacion_megadetallados",
        "sobre1_contenido",
        "sobre2_contenido",
        "sobre3_contenido",
    ]
    for k in watch_keys:
        if _has_banned_references(str(data.get(k, ""))):
            return True
    return False


# =========================================================
# HTTP helpers
# =========================================================
def _requests_get(url: str, timeout: int = 30) -> requests.Response:
    return requests.get(
        url,
        timeout=timeout,
        headers={"User-Agent": "Mozilla/5.0"},
        allow_redirects=True,
    )


def _absolute_url(base_url: str, href: str) -> str:
    href = (href or "").strip()
    if href.startswith("http://") or href.startswith("https://"):
        return href
    if href.startswith("//"):
        return "https:" + href
    if href.startswith("/"):
        m = re.match(r"^(https?://[^/]+)", base_url)
        if m:
            return m.group(1) + href
    return href


def _download_file(url: str, out_folder: str, filename: str) -> str:
    os.makedirs(out_folder, exist_ok=True)
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", filename)[:120]
    path = os.path.join(out_folder, safe)

    r = _requests_get(url, timeout=60)
    r.raise_for_status()

    with open(path, "wb") as f:
        f.write(r.content)
    return path


# =========================================================
# PLACSP: localizar PDF "Documento de Pliegos"
# =========================================================
def _find_pdf_links_in_row(row_tag, base_url: str) -> List[str]:
    candidates = []
    for a in row_tag.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        txt = a.get_text(" ", strip=True) or ""
        href_abs = _absolute_url(base_url, href)

        img = a.find("img")
        img_blob = ""
        if img is not None:
            img_blob = " ".join([
                str(img.get("alt", "")),
                str(img.get("title", "")),
                str(img.get("src", "")),
            ]).lower()

        blob = (href + " " + txt + " " + img_blob).lower()

        is_pdf = (
            ".pdf" in blob
            or "format=pdf" in blob
            or " pdf" in blob
            or "getdocumentbyidservlet" in blob
            or "getdocumentbyid" in blob
        )

        if is_pdf and href_abs.startswith("http"):
            candidates.append(href_abs)

    out, seen = [], set()
    for u in candidates:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def _find_placsp_pliego_pdf_url(tender_link: str) -> Optional[str]:
    r = _requests_get(tender_link, timeout=45)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    direct = []
    for a in soup.find_all("a", href=True):
        href = a.get("href") or ""
        if "GetDocumentByIdServlet" in href or "GetDocumentById" in href:
            u = _absolute_url(tender_link, href)
            if u.startswith("http"):
                direct.append(u)
    if direct:
        return direct[0]

    rows = soup.find_all("tr")
    best_links: List[str] = []
    for tr in rows:
        row_text = tr.get_text(" ", strip=True).lower()
        if "pliego" in row_text:
            links = _find_pdf_links_in_row(tr, tender_link)
            if links:
                links.sort(key=lambda u: (0 if "getdocumentbyidservlet" in u.lower() else 1))
                best_links.extend(links)

    if not best_links:
        all_links = []
        for a in soup.find_all("a", href=True):
            href = (a.get("href") or "").strip()
            txt = a.get_text(" ", strip=True) or ""
            u = _absolute_url(tender_link, href)
            if not u.startswith("http"):
                continue
            blob = (href + " " + txt).lower()
            if ".pdf" in blob or "getdocumentbyidservlet" in blob or "getdocumentbyid" in blob:
                all_links.append(u)
        all_links.sort(key=lambda u: (0 if "getdocumentbyidservlet" in u.lower() else 1))
        best_links = all_links

    seen, out = set(), []
    for u in best_links:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out[0] if out else None


# =========================================================
# Extraer links embebidos en PDF (para PCAP/PPT)
# =========================================================
def _extract_http_from_maybe_js(s: str) -> Optional[str]:
    if not s:
        return None
    s = str(s).strip()
    if s.startswith("http://") or s.startswith("https://"):
        return s
    m = re.search(r"(https?://[^\s\"')]+)", s)
    return m.group(1) if m else None


def _filespec_to_url(fspec) -> Optional[str]:
    if not fspec:
        return None
    if isinstance(fspec, str):
        return _extract_http_from_maybe_js(fspec)
    if isinstance(fspec, dict):
        for key in ("/UF", "/F"):
            val = fspec.get(key)
            if isinstance(val, str):
                u = _extract_http_from_maybe_js(val)
                if u:
                    return u
        if fspec.get("/FS") == "/URL":
            val = fspec.get("/F") or fspec.get("/UF")
            if isinstance(val, str):
                return _extract_http_from_maybe_js(val)
    return None


def _extract_pdf_links_actions(pdf_path: str) -> List[str]:
    urls: List[str] = []
    try:
        try:
            from pypdf import PdfReader  # type: ignore
        except Exception:
            from PyPDF2 import PdfReader  # type: ignore

        reader = PdfReader(pdf_path)
        for page in reader.pages:
            annots = page.get("/Annots", []) or []
            for a in annots:
                try:
                    obj = a.get_object()
                except Exception:
                    obj = a
                if not obj:
                    continue

                action = obj.get("/A") or obj.get("/AA") or None
                if not action:
                    continue

                try:
                    if hasattr(action, "get_object"):
                        action = action.get_object()
                except Exception:
                    pass

                if not isinstance(action, dict):
                    continue

                uri = action.get("/URI")
                if isinstance(uri, str) and uri:
                    u = _extract_http_from_maybe_js(uri)
                    if u:
                        urls.append(u)
                    continue

                s_type = action.get("/S")
                if s_type in ("/GoToR", "/Launch", "/GoToE", "/GoTo"):
                    fspec = action.get("/F")
                    u = _filespec_to_url(fspec)
                    if u:
                        urls.append(u)
                        continue

                fspec = action.get("/F")
                u = _filespec_to_url(fspec)
                if u:
                    urls.append(u)
                    continue

    except Exception:
        return []

    out, seen = [], set()
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def _extract_pdf_urls_fallback_text(pdf_path: str) -> List[str]:
    try:
        raw = extract_text(pdf_path) or ""
    except Exception:
        raw = ""
    raw = raw.replace("\r", "\n")
    urls = re.findall(r"https?://[^\s)\]]+", raw)
    cleaned = [u.strip().rstrip(".,;:") for u in urls]
    out, seen = [], set()
    for u in cleaned:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def _extract_candidate_doc_links_from_html(tender_link: str) -> List[str]:
    try:
        r = _requests_get(tender_link, timeout=45)
        r.raise_for_status()
        html = r.text or ""
        soup = BeautifulSoup(html, "html.parser")
    except Exception:
        return []

    urls: List[str] = []
    for a in soup.find_all("a", href=True):
        href = a.get("href") or ""
        if "GetDocumentByIdServlet" in href or "GetDocumentById" in href:
            u = _absolute_url(tender_link, href)
            if u.startswith("http"):
                urls.append(u)

    for m in re.findall(r"(https?://[^\s\"']+GetDocumentByIdServlet[^\s\"']+)", html):
        urls.append(m)

    out, seen = [], set()
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def _guess_doc_type_from_text(pdf_path: str) -> Optional[str]:
    try:
        text = extract_text(pdf_path) or ""
    except Exception:
        return None
    t = re.sub(r"\s+", " ", text).lower()

    if "cláusulas administrativas" in t or "clausulas administrativas" in t:
        return "pcap"
    if "pliego de cláusulas" in t or "pliego de clausulas" in t:
        return "pcap"
    if "prescripciones técnicas" in t or "prescripciones tecnicas" in t:
        return "ppt"
    if "pliego de prescripciones" in t:
        return "ppt"
    return None


def _classify_pliego_urls_by_name(urls: List[str]) -> Tuple[Optional[str], Optional[str]]:
    pcap = None
    ppt = None

    def score_pcap(u: str) -> int:
        s = u.lower()
        sc = 0
        if "pcap" in s: sc += 25
        if "clausul" in s or "cláusul" in s: sc += 18
        if "administr" in s: sc += 14
        if "prescrip" in s or "tecnic" in s or "ppt" in s: sc -= 10
        return sc

    def score_ppt(u: str) -> int:
        s = u.lower()
        sc = 0
        if "ppt" in s: sc += 25
        if "prescrip" in s: sc += 18
        if "tecnic" in s or "técnic" in s: sc += 14
        if "clausul" in s or "administr" in s or "pcap" in s: sc -= 10
        return sc

    filtered = [u for u in urls if u.startswith("http")]
    if not filtered:
        return None, None

    best_pcap = sorted(filtered, key=score_pcap, reverse=True)[0]
    best_ppt = sorted(filtered, key=score_ppt, reverse=True)[0]

    if score_pcap(best_pcap) > 0:
        pcap = best_pcap
    if score_ppt(best_ppt) > 0:
        ppt = best_ppt

    if pcap and ppt and pcap == ppt:
        ppt = None
    return pcap, ppt


def download_pliegos_from_tender_page(
    tender_link: str,
    cache_folder: str,
    tender_id: str
) -> Tuple[Optional[str], Optional[str], str]:
    os.makedirs(cache_folder, exist_ok=True)

    pliego_pdf_url = _find_placsp_pliego_pdf_url(tender_link)
    if not pliego_pdf_url:
        return None, None, "No he encontrado el PDF de 'Pliego' en 'Anuncios y Documentos'."

    intermediate_pdf = _download_file(pliego_pdf_url, cache_folder, f"{tender_id}_PLIEGO_ANUNCIO.pdf")

    urls = _extract_pdf_links_actions(intermediate_pdf)
    source_used = "anotaciones PDF (robusto: URI/GoToR/Launch)"
    if not urls:
        urls = _extract_pdf_urls_fallback_text(intermediate_pdf)
        source_used = "texto PDF (fallback regex)"
    if not urls:
        urls = _extract_candidate_doc_links_from_html(tender_link)
        source_used = "HTML expediente (fallback GetDocumentByIdServlet)"

    info_parts = [
        f"PDF 'Pliego' localizado: {pliego_pdf_url}",
        f"Extracción de enlaces: {source_used}",
        f"URIs detectadas: {len(urls)}",
    ]

    if not urls:
        info_parts.append("❌ No se han encontrado enlaces a documentos (PCAP/PPT).")
        return None, None, " | ".join(info_parts)

    pcap_url, ppt_url = _classify_pliego_urls_by_name(urls)

    pcap_path = None
    ppt_path = None

    def _dl_candidate(u: str, filename: str) -> Optional[str]:
        try:
            return _download_file(u, cache_folder, filename)
        except Exception:
            return None

    if pcap_url:
        p = _dl_candidate(pcap_url, f"{tender_id}_PCAP.pdf")
        if p:
            pcap_path = p
            info_parts.append("PCAP descargado.")
    if ppt_url:
        p = _dl_candidate(ppt_url, f"{tender_id}_PPT.pdf")
        if p:
            ppt_path = p
            info_parts.append("PPT descargado.")

    if not pcap_path or not ppt_path:
        candidates_to_try = urls[:12]
        for idx, u in enumerate(candidates_to_try, start=1):
            if (pcap_url and u == pcap_url) or (ppt_url and u == ppt_url):
                continue
            cand = _dl_candidate(u, f"{tender_id}_CAND_{idx}.pdf")
            if not cand:
                continue
            doc_type = _guess_doc_type_from_text(cand)
            if doc_type == "pcap" and not pcap_path:
                final_path = os.path.join(cache_folder, f"{tender_id}_PCAP.pdf")
                try:
                    os.replace(cand, final_path)
                    pcap_path = final_path
                except Exception:
                    pcap_path = cand
            if doc_type == "ppt" and not ppt_path:
                final_path = os.path.join(cache_folder, f"{tender_id}_PPT.pdf")
                try:
                    os.replace(cand, final_path)
                    ppt_path = final_path
                except Exception:
                    ppt_path = cand
            if pcap_path and ppt_path:
                break

    if not pcap_path:
        info_parts.append("⚠️ No se ha podido identificar PCAP.")
    if not ppt_path:
        info_parts.append("⚠️ No se ha podido identificar PPT.")

    return pcap_path, ppt_path, " | ".join(info_parts)


# =========================================================
# PDF -> Texto
# =========================================================
def pdf_to_text_keep_pages(pdf_path: str, max_chars: int = 850000) -> str:
    raw = extract_text(pdf_path) or ""
    raw = raw.replace("\r", "")
    if len(raw) > max_chars:
        raw = raw[:max_chars]
    return raw


def split_pages(raw_text: str) -> List[str]:
    parts = re.split(r"\f+", raw_text or "")
    out = []
    for p in parts:
        p2 = re.sub(r"\s+", " ", p).strip()
        if p2:
            out.append(p2)
    return out if out else []


# =========================================================
# Selección de páginas (MEJORADA)
# =========================================================
_KEYWORDS_WEIGHTED = [
    ("expediente", 2), ("objeto", 3), ("código cpv", 5), ("cpv", 3),
    ("presupuesto", 3), ("presupuesto base", 4), ("valor estimado", 4),
    ("duración", 2), ("duracion", 2), ("plazo de ejecución", 4), ("plazo de ejecucion", 4),
    ("fecha límite", 4), ("fecha limite", 4), ("presentación de proposiciones", 4),
    ("presentacion de proposiciones", 4), ("hora", 1), ("lugar", 2),

    ("visita", 6), ("visita técnica", 8), ("visita tecnica", 8),
    ("reconocimiento", 6), ("inspección", 5), ("inspeccion", 5),

    ("solvencia", 6), ("solvencia económica", 7), ("solvencia economica", 7),
    ("solvencia técnica", 8), ("solvencia tecnica", 8),
    ("volumen de negocio", 6), ("facturación", 6), ("facturacion", 6),
    ("seguro", 5), ("responsabilidad civil", 6),

    ("adscripción", 9), ("adscripcion", 9),
    ("medios personales", 10), ("medios materiales", 10),
    ("equipo", 6), ("perfil", 5),
    ("currículum", 6), ("curriculum", 6),

    ("criterios de adjudicación", 12), ("criterios de adjudicacion", 12),
    ("puntuación", 8), ("puntuacion", 8),
    ("fórmula", 8), ("formula", 8),
    ("juicio de valor", 7), ("automático", 7), ("automatico", 7),
    ("ofertas desproporcionadas", 7), ("temerarias", 6),

    ("sobre", 10), ("sobres", 10),
    ("archivo electrónico", 10), ("archivo electronico", 10),
    ("documentación", 9), ("documentacion", 9),
    ("deuc", 8), ("declaración responsable", 7), ("declaracion responsable", 7),
    ("anexo", 7), ("modelo", 6),

    # ✅ Refuerzo anexos / cuadro resumen
    ("anexo i", 12), ("anexo 1", 12),
    ("cuadro resumen", 10),
    ("resumen", 4),

    ("penalidades", 4),
]


def _page_score(pl: str) -> int:
    score = 0
    for kw, w in _KEYWORDS_WEIGHTED:
        if kw in pl:
            score += w
    if re.search(r"\b\d{8}-\d\b", pl):
        score += 6
    if re.search(r"\d{1,2}/\d{1,2}/\d{4}", pl) or "€" in pl or "eur" in pl:
        score += 2
    return score


def _select_relevant_pages(pages: List[str], max_pages: int = 34) -> List[int]:
    if not pages:
        return []
    scored = []
    for i, p in enumerate(pages):
        pl = (p or "").lower()
        sc = _page_score(pl)
        if sc > 0:
            scored.append((sc, i))

    if not scored:
        return list(range(min(max_pages, len(pages))))

    scored.sort(key=lambda x: (-x[0], x[1]))
    top_idxs = [i for _, i in scored[:max_pages]]

    expanded = set()
    for idx in top_idxs:
        expanded.add(idx)
        for j in range(idx - 2, idx + 3):
            if 0 <= j < len(pages):
                expanded.add(j)

    expanded.add(0)
    out = sorted(expanded)
    return out[:max_pages]


def _pack_pages(pages: List[str], picked: List[int], label: str, max_len: int = 6500) -> str:
    blocks = []
    for idx in picked:
        pno = idx + 1
        blocks.append(f"[{label} PÁG {pno}]\n{pages[idx][:max_len]}")
    return "\n\n".join(blocks)


# =========================================================
# LLM helpers: fallbacks + heurísticas
# =========================================================
def _extract_cpv_regex(full_text: str) -> str:
    if not full_text:
        return ""
    cpvs = re.findall(r"\b\d{8}-\d\b", full_text)
    out, seen = [], set()
    for c in cpvs:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return ", ".join(out)


def _infer_visita_tecnica(full_text: str, current: str) -> str:
    cur = _clean_inline(current)
    if cur.lower() in ("si", "sí", "no"):
        return "Si" if cur.lower() in ("si", "sí") else "No"

    t = (full_text or "").lower()
    if any(k in t for k in ["visita", "visita técnica", "visita tecnica", "reconocimiento", "inspección", "inspeccion"]):
        if "no obligatoria" in t or "no es obligatoria" in t or "no será obligatoria" in t:
            return "No"
        if "obligatoria" in t or "obligatorio" in t:
            return "Si"
        return ""
    return "No"



def _infer_criterios_100pct(full_text: str, current: str) -> str:
    """
    Si el modelo dejó criterios vacíos pero el pliego indica claramente un criterio único al 100%,
    rellenamos una línea mínima para no dejar la celda en blanco.
    """
    cur = _clean_inline(current)
    if cur:
        return cur

    t = _normalize(full_text or "")

    # Detectores de "criterio único" + palabras clave
    unico = bool(re.search(r"\b(criterio\s+unico|unico\s+criterio|criterio\s+exclusivo)\b", t))
    tiene_100 = ("100%" in (full_text or "")) or bool(re.search(r"\b100\b", t))

    if not unico and not tiene_100:
        return ""

    # Precio / oferta económica
    if any(k in t for k in ["precio", "oferta economica", "oferta económica", "importe", "coste", "costo"]):
        if unico or ("100%" in (full_text or "")) or re.search(r"\b100\s*(puntos|pts)\b", t):
            return " - Precio / oferta económica: 100 puntos"

    # Técnico / calidad / memoria
    if any(k in t for k in ["memoria tecnica", "memoria técnica", "calidad", "criterios tecnicos", "criterios técnicos", "juicio de valor"]):
        if unico or ("100%" in (full_text or "")) or re.search(r"\b100\s*(puntos|pts)\b", t):
            return " - Criterio técnico / calidad: 100 puntos"

    return ""
def _is_too_short(s: str, min_chars: int) -> bool:
    return len(_clean_inline(s)) < min_chars


def _pick_pages_by_terms(pages: List[str], terms: List[str], max_take: int = 18) -> List[int]:
    scored = []
    for i, p in enumerate(pages):
        pl = (p or "").lower()
        sc = 0
        for t in terms:
            if t in pl:
                sc += 1
        if sc > 0:
            scored.append((sc, i))
    scored.sort(key=lambda x: (-x[0], x[1]))
    idxs = [i for _, i in scored[:max_take]]

    out = set()
    for idx in idxs:
        for j in range(idx - 1, idx + 2):
            if 0 <= j < len(pages):
                out.add(j)
    out.add(0)
    return sorted(out)[:max_take]


def _fields_suspicious_or_empty(data: Dict) -> bool:
    """
    Detecta "vacíos sospechosos": demasiado corto o con referencias prohibidas en campos críticos.
    """
    # mínimos orientativos (ajustables)
    checks = [
        ("objeto_detallado", 300),
        ("solvencia_tecnica_detallada", 450),
        ("medios_humanos_materiales_detallados", 650),
        ("criterios_adjudicacion_megadetallados", 950),
        ("sobre1_contenido", 220),
    ]

    for k, minc in checks:
        v = str(data.get(k, "") or "")
        if _has_banned_references(v):
            return True
        if _is_too_short(v, minc):
            return True

    # sobres 2/3: si hay título pero contenido corto → sospechoso
    for n in (2, 3):
        titulo = str(data.get(f"sobre{n}_titulo", "") or "")
        cont = str(data.get(f"sobre{n}_contenido", "") or "")
        if _clean_inline(titulo) and (_has_banned_references(cont) or _is_too_short(cont, 180)):
            return True

    return False


# =========================================================
# LLM: devolver JSON mapeable al Excel (GPT-5.2 Responses)
#   - SIN response_format (SDK antiguo)
#   - max_output_tokens=6500
# =========================================================
def llm_generate_excel_fields(
    tender_title: str,
    tender_link: str,
    pcap_text: str,
    ppt_text: str,
    progress_cb: Optional[Callable[[float, str], None]] = None
) -> Dict:
    client = _get_client()
    if progress_cb:
        try:
            progress_cb(0.35, "Preparando contexto IA…")
        except Exception:
            pass

    pcap_pages = split_pages(pcap_text)
    ppt_pages = split_pages(ppt_text) if ppt_text else []

    pcap_pick = _select_relevant_pages(pcap_pages, max_pages=34)
    pcap_ctx = _pack_pages(pcap_pages, pcap_pick, "PCAP", max_len=6500)

    ppt_ctx = ""
    if ppt_pages:
        ppt_pick = _select_relevant_pages(ppt_pages, max_pages=16)
        ppt_ctx = _pack_pages(ppt_pages, ppt_pick, "PPT", max_len=5200)

    prompt = f"""
Eres un TÉCNICO SENIOR DE LICITACIONES. Extrae información del PCAP y (si existe) PPT para rellenar una plantilla Excel.
Tu objetivo es que alguien pueda TRABAJAR la licitación usando SOLO el Excel.

REGLAS:
0) Devuelve TODO en CASTELLANO (español), aunque el pliego esté en otro idioma. Traduce al castellano cualquier fragmento que extraigas.
1) NO inventes. Si no aparece explícito: "".
2) PROHIBIDO lenguaje genérico. Si no hay detalle concreto, pon "".
3) PROHIBIDO referenciar sin volcar el contenido. Ejemplos prohibidos:
   - "conforme al Anexo I del PCAP", "según PCAP", "ver cláusula X", "ver cuadro resumen"
   Si mencionas "Anexo/Cláusula/Apartado", DEBES incluir el contenido/valores concretos aquí en el Excel.
4) El Excel debe ser AUTOSUFICIENTE: quien lo lea NO debe abrir los pliegos.
5) Muy detallado en: objeto, solvencia técnica, medios, criterios, sobres.
5.1) OBJETO DEL CONTRATO:
   - Debe ser útil para preparar la oferta, pero resumido.
   - Máximo aproximado de 5 líneas.
   - Explica qué hay que hacer, alcance, entregables y finalidad, sin copiar párrafos larguísimos del pliego.

6) FORMATO DE LISTAS EN TODO EL EXCEL:
   - SIEMPRE que listes elementos, cada elemento debe empezar por "- " y terminar con un salto de línea real dentro de la celda.
   - PROHIBIDO poner varios elementos en una sola línea separados por " - ", ";", comas o texto corrido.
   - Cada guion debe ocupar su propia línea. Si hay 5 elementos, deben verse 5 líneas distintas en Excel.
   - Esto aplica a cualquier campo con listados, especialmente medios, criterios y sobres.

7) MEDIOS HUMANOS Y MATERIALES:
   - Lista con "- " SOLO los PERFILES solicitados (un perfil por línea y con salto de línea después de cada uno).
   - Si piden experiencia mínima de un perfil, inclúyela en la misma línea del perfil.
   - El resto de información que no sea un perfil obligatorio debe ir en texto normal SIN guiones, en párrafos separados.
   - Si el pliego indica EXPRESAMENTE la titulación/profesión que puede cumplir cada función/perfil, indícalo entre paréntesis.

8) CRITERIOS DE ADJUDICACIÓN:
   - Revísalos de forma exhaustiva en TODAS las partes del pliego donde aparezcan: cuadro resumen, PCAP, PPT, anexos citados dentro del texto extraído, tablas y apartados de fórmulas.
   - Lista SIEMPRE con "- " y un criterio por línea, con salto de línea real tras cada línea.
   - Indica SIEMPRE la puntuación exacta en puntos de cada criterio.
   - Si hay subcriterios, desglósalos todos indicando los puntos de cada subapartado, por pequeños que sean.
   - Si hay fórmulas, umbrales, mejoras, bolsas de horas, ampliaciones, reducciones de plazo, metodologías, memorias, equipo adscrito, certificados o cualquier otro subcriterio, inclúyelo con su puntuación concreta.
   - No pongas resúmenes vagos del tipo "100% precio" si realmente hay más desglose o reglas; detállalo completo.
   - Si el pliego NO indica puntuación/reparto, deja "criterios_adjudicacion_megadetallados" como "" (vacío), EXCEPTO si detectas claramente que el criterio es único al 100% (por ejemplo: precio 100 puntos / 100%); en ese caso, indícalo.

9) SOBRES:
   - Lista exacta SOLO de documentación a subir para PRESENTAR la oferta (fase de licitación).
   - NO incluyas documentación que se aporta tras la adjudicación/por el adjudicatario.
   - NO incluyas acciones o instrucciones que no sean documentos (ej.: "firma electrónica del representante").
   - 1 documento por línea con "- ".

9) IMPORTE:
   - "importe_licitacion" debe ser SIEMPRE SIN impuestos (sin IVA).
   - Si el pliego solo da importe con IVA, calcula y devuelve el importe sin IVA.

DEVUELVE ESTE JSON EXACTO (mismas claves):
{{
  "entidad_adjudicadora": "",
  "importe_licitacion": "",
  "fecha_limite": "",
  "hora_limite": "",
  "plazo_ejecucion": "",
  "titulo": "{tender_title}",
  "expediente": "",
  "cpv": "",
  "visita_tecnica": "Si/No o ''",
  "lugar_prestacion": "",
  "objeto_detallado": "",
  "solvencia_tecnica_detallada": "",
  "medios_humanos_materiales_detallados": "",
  "criterios_adjudicacion_megadetallados": "",

  "sobre1_titulo": "",
  "sobre1_contenido": "",
  "sobre1_pagina": "",

  "sobre2_titulo": "",
  "sobre2_contenido": "",
  "sobre2_pagina": "",

  "sobre3_titulo": "",
  "sobre3_contenido": "",
  "sobre3_pagina": ""
}}

CONTEXTO PCAP:
\"\"\"{pcap_ctx}\"\"\"

CONTEXTO PPT (si hay):
\"\"\"{ppt_ctx}\"\"\"
"""

    if progress_cb:
        try:
            progress_cb(0.45, "IA: extracción inicial…")
        except Exception:
            pass

    r = client.responses.create(
        model=os.getenv("OPENAI_MODEL", "gpt-5.4"),
        input=prompt,
        max_output_tokens=6500,
    )

    content = _responses_output_text(r)
    if not content:
        raise RuntimeError("OpenAI devolvió respuesta vacía (sin texto).")

    data = _parse_json_or_raise(content)

    def _yn(v: str) -> str:
        v = _clean_inline(v).lower()
        if not v:
            return ""
        if v in ("si", "sí", "yes"):
            return "Si"
        if v in ("no",):
            return "No"
        if "si" in v or "sí" in v:
            return "Si"
        if "no" in v:
            return "No"
        return ""

    keys = [
        "entidad_adjudicadora", "importe_licitacion", "fecha_limite", "hora_limite", "plazo_ejecucion",
        "titulo", "expediente", "cpv", "visita_tecnica", "lugar_prestacion",
        "objeto_detallado", "solvencia_tecnica_detallada", "medios_humanos_materiales_detallados",
        "criterios_adjudicacion_megadetallados",
        "sobre1_titulo", "sobre1_contenido", "sobre1_pagina",
        "sobre2_titulo", "sobre2_contenido", "sobre2_pagina",
        "sobre3_titulo", "sobre3_contenido", "sobre3_pagina",
    ]
    multiline_keys = {
        "objeto_detallado", "solvencia_tecnica_detallada", "medios_humanos_materiales_detallados",
        "criterios_adjudicacion_megadetallados", "sobre1_contenido", "sobre2_contenido", "sobre3_contenido"
    }
    for k in keys:
        if k not in data or data[k] is None:
            data[k] = ""
        else:
            data[k] = _clean_keep_newlines(data[k]) if k in multiline_keys else _clean_inline(data[k])

    data["visita_tecnica"] = _yn(data.get("visita_tecnica", ""))
    data["visita_tecnica"] = _infer_visita_tecnica(pcap_text + "\n" + (ppt_text or ""), data["visita_tecnica"])
    if not data.get("cpv"):
        data["cpv"] = _extract_cpv_regex(pcap_text)

    # =========================================================
    # 2ª pasada (ya la tenías): refuerzo de MEDIOS/CRITERIOS/SOBRES si salen pobres
    # =========================================================
    need_boost = (
        _is_too_short(data.get("medios_humanos_materiales_detallados", ""), 600) or
        _is_too_short(data.get("criterios_adjudicacion_megadetallados", ""), 900) or
        _is_too_short(data.get("sobre1_contenido", ""), 220) or
        (_is_too_short(data.get("sobre2_contenido", ""), 180) and data.get("sobre2_titulo")) or
        (_is_too_short(data.get("sobre3_contenido", ""), 180) and data.get("sobre3_titulo"))
    )

    if need_boost:
        focus_terms = [
            "medios personales", "medios materiales", "adscripción", "adscripcion", "perfil", "currículum", "curriculum",
            "criterios de adjudicación", "criterios de adjudicacion", "puntuación", "puntuacion", "fórmula", "formula",
            "sobre", "archivo electrónico", "archivo electronico", "documentación", "documentacion", "deuc", "modelo", "anexo",
            "anexo i", "anexo 1", "cuadro resumen",
        ]

        pcap_pick2 = _pick_pages_by_terms(pcap_pages, focus_terms, max_take=18)
        pcap_ctx2 = _pack_pages(pcap_pages, pcap_pick2, "PCAP", max_len=8000)

        prompt2 = f"""
Mejora este JSON SIN inventar. Completa MEDIOS, CRITERIOS y SOBRES con listas exactas y detalles operativos.
REGLAS:
0) Devuelve TODO en CASTELLANO (español), aunque el pliego esté en otro idioma. Traduce al castellano cualquier fragmento que extraigas.
- Devuelve SOLO JSON.
- PROHIBIDO: "según el anexo", "conforme al PCAP", "ver cláusula/apartado".
  Sustituye referencias por contenido concreto (requisitos, importes, umbrales, fórmulas, documentación exacta).
- Sobres: 1 documento por línea con "- " y salto de línea real después de cada documento.
- Criterios: no dejes el campo en blanco si detectas claramente que es un criterio único al 100% (p. ej., precio 100 puntos / 100%), pero SOLO si de verdad no hay más criterios.
- Medios: lista con "- " SOLO los perfiles; cada perfil en una línea distinta. Lo demás en texto normal sin guiones.
- Medios: si el pliego indica expresamente titulación/profesión para un perfil, ponlo entre paréntesis; si no, no lo pongas.
- Criterios: indica SIEMPRE la puntuación exacta (puntos) y el reparto por subcriterios; no aceptes una salida simplificada tipo "precio 100 puntos" si el pliego contiene más criterios técnicos o subapartados; si no hay puntos, deja el campo vacío ("").
- Sobres: SOLO documentación para presentar oferta (no post-adjudicación) y SOLO documentos (no acciones como firma electrónica).
- Importe: "importe_licitacion" siempre sin IVA.

JSON ACTUAL:
{json.dumps(data, ensure_ascii=False)}

CONTEXTO PCAP ENFOCADO:
\"\"\"{pcap_ctx2}\"\"\"
"""

        if progress_cb:
            try:
                progress_cb(0.62, "IA: refuerzo de medios/criterios/sobres…")
            except Exception:
                pass

        r2 = client.responses.create(
            model=os.getenv("OPENAI_MODEL", "gpt-5.4"),
            input=prompt2,
            max_output_tokens=6500,
        )

        content2 = _responses_output_text(r2)
        if content2:
            data2 = _parse_json_or_raise(content2)
            if isinstance(data2, dict):
                for k in keys:
                    if k in data2 and data2[k] is not None:
                        data[k] = _clean_keep_newlines(data2[k]) if k in multiline_keys else _clean_inline(data2[k])

                data["visita_tecnica"] = _infer_visita_tecnica(pcap_text + "\n" + (ppt_text or ""), data.get("visita_tecnica", ""))
                if not data.get("cpv"):
                    data["cpv"] = _extract_cpv_regex(pcap_text)

    # =========================================================
    # 3ª pasada: eliminar referencias tipo "según anexo/PCAP" y volcar contenido
    # (esto soluciona lo que pediste y evita frases estilo "conforme al Anexo I")
    # =========================================================
    if _fields_need_dereference(data):
        prompt3 = f"""
Eres un TÉCNICO SENIOR DE LICITACIONES. Tu tarea es CORREGIR este JSON para que el Excel sea AUTOSUFICIENTE.
IMPORTANTE: Devuelve TODO en castellano (español), aunque el pliego esté en otro idioma. Traduce cualquier fragmento.

PROHIBIDO:
- Frases como "conforme al anexo", "según PCAP", "ver cláusula", "ver apartado", "ver cuadro resumen".
- Referencias sin volcar el contenido.

OBLIGATORIO:
- Sustituye cualquier referencia por el CONTENIDO/VALORES concretos (requisitos, importes, plazos, criterios,
- Criterios: no dejes en blanco si detectas claramente criterio único al 100% (p.ej. precio 100 puntos / 100%).
- Medios: lista con "- " SOLO los perfiles; el resto en texto normal sin guiones.
  documentación por sobre, umbrales, fórmulas, penalidades, etc.).
- Mantén el formato de SOBRES: 1 documento por línea con "- ".
- En SOBRES: SOLO documentación para presentar oferta (no post-adjudicación) y SOLO documentos (no acciones).
- En MEDIOS: profesión/titulación entre paréntesis SOLO si el pliego lo indica explícitamente.
- En CRITERIOS: incluye SIEMPRE puntos y reparto; si no hay puntos, deja el campo vacío ("").
- En IMPORTE: "importe_licitacion" siempre sin IVA.
- Si NO encuentras un dato explícito, deja "".

Devuelve SOLO JSON con las MISMAS CLAVES (no añadas ni quites claves).

JSON A CORREGIR:
{json.dumps(data, ensure_ascii=False)}

CONTEXTO PCAP:
\"\"\"{pcap_ctx}\"\"\"

CONTEXTO PPT (si hay):
\"\"\"{ppt_ctx}\"\"\"
"""

        if progress_cb:
            try:
                progress_cb(0.74, "IA: eliminando referencias a anexos/PCAP…")
            except Exception:
                pass

        r3 = client.responses.create(
            model=os.getenv("OPENAI_MODEL", "gpt-5.4"),
            input=prompt3,
            max_output_tokens=6500,
        )

        content3 = _responses_output_text(r3)
        if content3:
            data3 = _parse_json_or_raise(content3)
            if isinstance(data3, dict):
                for k in keys:
                    if k in data3 and data3[k] is not None:
                        data[k] = _clean_keep_newlines(data3[k]) if k in multiline_keys else _clean_inline(data3[k])

                data["visita_tecnica"] = _infer_visita_tecnica(pcap_text + "\n" + (ppt_text or ""), data.get("visita_tecnica", ""))
                if not data.get("cpv"):
                    data["cpv"] = _extract_cpv_regex(pcap_text)

    # =========================================================
    # 4ª pasada: "vacíos sospechosos" (lo que me pediste)
    # Si sigue corto o referencial tras las pasadas anteriores, hacemos un prompt quirúrgico
    # =========================================================
    if _fields_suspicious_or_empty(data):
        focus_terms3 = [
            # objetivo / alcance / entregables
            "objeto", "alcance", "prestación", "entregable", "entregables", "memoria", "planos", "presupuesto", "mediciones",
            # solvencia / medios
            "solvencia", "clasificación", "clasificacion", "experiencia", "volumen de negocio", "facturación", "adscripción",
            "adscripcion", "medios personales", "medios materiales", "equipo", "perfil", "titulación", "titulacion",
            # criterios / formulas
            "criterios de adjudicación", "criterios de adjudicacion", "puntuación", "puntuacion", "fórmula", "formula",
            "juicio de valor", "automático", "automatico", "umbral", "baja temeraria", "temeraria",
            # sobres / doc
            "sobre", "archivo electrónico", "archivo electronico", "documentación", "documentacion", "deuc", "anexo", "modelo",
            "anexo i", "anexo 1", "cuadro resumen",
            # penalidades / garantías
            "garantía", "garantia", "penalidades", "penalización", "penalizacion",
        ]
        pcap_pick3 = _pick_pages_by_terms(pcap_pages, focus_terms3, max_take=22)
        pcap_ctx3 = _pack_pages(pcap_pages, pcap_pick3, "PCAP", max_len=9000)

        prompt4 = f"""
Vas a COMPLETAR los campos críticos para que el Excel sea AUTOSUFICIENTE y OPERATIVO.
IMPORTANTE: Devuelve TODO en castellano (español), aunque el pliego esté en otro idioma. Traduce cualquier fragmento.

PROHIBIDO:
- Cualquier frase referencial: "según el anexo", "conforme al PCAP", "ver cláusula/apartado/cuadro".
- Texto genérico.

OBLIGATORIO:
- Volcar el contenido concreto: requisitos, umbrales, cifras, plazos, documentación exacta, fórmulas, criterios y reparto.
- Objeto: resumen útil y operativo, de unas 5 líneas máximas aprox., sin copiar bloques larguísimos del pliego.
- Criterios: revísalos a fondo en todas las partes del pliego; no dejes en blanco si detectas claramente criterio único al 100% (p.ej. precio 100 puntos / 100%), pero ojo, solo si realmente ves que el criterio es 100% precio porque es raro que eso sea así, casi siempre los criterios se dividen en diferentes apartados y el precio solo es una parte.
- Criterios: cada criterio y subcriterio debe ir con "- " y salto de línea; indica los puntos concretos de cada cosa, por pequeña que sea.
- Medios: lista con "- " SOLO los perfiles e inserta un salto de línea después de cada perfil como es excel se hace con alt + enter así visualmente queda mejor; si piden años de experiencia también inclúyelos en la misma línea. El resto en texto normal sin guiones. No me cuentes mucha historia aquí, me interesa saber qué perfiles pide el pliego para cumplir.
- Cualquier listado en cualquier campo debe ir siempre con "- " y salto de línea después de cada elemento; nunca dejes varios ítems seguidos en la misma línea.
- SOBRES: 1 documento por línea con "- ".
- En SOBRES: SOLO documentación para presentar oferta (no post-adjudicación) y SOLO documentos (no acciones).
- En MEDIOS: profesión/titulación entre paréntesis SOLO si el pliego lo indica explícitamente.
- En CRITERIOS: incluye SIEMPRE puntos y reparto; si no hay puntos, deja el campo vacío ("").
- En IMPORTE: "importe_licitacion" siempre sin IVA.
- Si un dato NO aparece explícito, deja "".

Devuelve SOLO JSON con las MISMAS CLAVES.

JSON ACTUAL:
{json.dumps(data, ensure_ascii=False)}

CONTEXTO PCAP HIPER-ENFOCADO:
\"\"\"{pcap_ctx3}\"\"\"
"""

        if progress_cb:
            try:
                progress_cb(0.86, "IA: completando vacíos críticos…")
            except Exception:
                pass

        r4 = client.responses.create(
            model=os.getenv("OPENAI_MODEL", "gpt-5.4"),
            input=prompt4,
            max_output_tokens=6500,
        )

        content4 = _responses_output_text(r4)
        if content4:
            data4 = _parse_json_or_raise(content4)
            if isinstance(data4, dict):
                for k in keys:
                    if k in data4 and data4[k] is not None:
                        data[k] = _clean_keep_newlines(data4[k]) if k in multiline_keys else _clean_inline(data4[k])

                data["visita_tecnica"] = _infer_visita_tecnica(pcap_text + "\n" + (ppt_text or ""), data.get("visita_tecnica", ""))
                if not data.get("cpv"):
                    data["cpv"] = _extract_cpv_regex(pcap_text)

    # Heurística final: criterio único al 100% para no dejar la celda vacía
    data["criterios_adjudicacion_megadetallados"] = _infer_criterios_100pct(
        pcap_text + "\n" + (ppt_text or ""),
        data.get("criterios_adjudicacion_megadetallados", "")
    )

    return data


# =========================================================
# Excel writing helpers
# =========================================================

# =========================================================
# Excel prettify helpers (multilínea con guiones)
# =========================================================
def _split_to_items(raw: str) -> List[str]:
    s = _clean_keep_newlines(raw)
    if not s:
        return []

    s = re.sub(r"(?<!^)\s+(?=[-•·●▪]\s+)", "\n", s)
    s = re.sub(r"([.;:])\s+(?=[-•·●▪]\s+)", r"\1\n", s)
    s = s.replace("•", "\n- ").replace("·", "\n- ").replace("●", "\n- ").replace("▪", "\n- ")
    s = re.sub(r"\n{2,}", "\n", s)

    items = [ln.strip() for ln in s.split("\n") if ln.strip()]
    if len(items) >= 2:
        return items

    one = items[0] if items else s
    if ";" in one:
        parts = [p.strip() for p in one.split(";") if p.strip()]
        if len(parts) >= 2:
            return parts
    if re.search(r"\s-\s", one):
        parts = [p.strip() for p in re.split(r"\s-\s", one) if p.strip()]
        if len(parts) >= 2:
            return parts
    return [one]


def _as_bulleted_multiline(raw: str) -> str:
    items = _split_to_items(raw)
    if not items:
        return ""

    cleaned: List[str] = []
    for it in items:
        t = re.sub(r"^(?:[-•·●▪]+)\s*", "", it.strip())
        if t:
            cleaned.append(f"- {t}")
    return "\n".join(cleaned)


def _as_mixed_multiline(raw: str) -> str:
    s = _clean_keep_newlines(raw)
    if not s:
        return ""

    s = re.sub(r"(?<!^)\s+(?=[-•·●▪]\s+)", "\n", s)
    s = re.sub(r"([.;:])\s+(?=[-•·●▪]\s+)", r"\1\n", s)
    s = re.sub(r"\s+(?=(?:Medios materiales|Materiales exigidos|Medios materiales exigidos|Equipos mínimos|Equipamiento mínimo)\b)", "\n", s, flags=re.IGNORECASE)
    s = re.sub(r"\n{3,}", "\n\n", s)

    out_lines: List[str] = []
    for ln in s.split("\n"):
        t = ln.strip()
        if not t:
            if out_lines and out_lines[-1] != "":
                out_lines.append("")
            continue

        if re.match(r"^[-•·●▪]\s+", t):
            out_lines.append(f"- {re.sub(r"^[-•·●▪]+\s*", "", t).strip()}")
            continue

        if re.search(r"\s[-•·●▪]\s+", t):
            parts = [p.strip() for p in re.split(r"(?=\s[-•·●▪]\s+)|(?=^[-•·●▪]\s+)", t) if p.strip()]
            used = False
            for part in parts:
                if re.match(r"^[-•·●▪]\s+", part):
                    out_lines.append(f"- {re.sub(r"^[-•·●▪]+\s*", "", part).strip()}")
                    used = True
                else:
                    out_lines.append(part)
                    used = True
            if used:
                continue

        out_lines.append(t)

    while out_lines and out_lines[-1] == "":
        out_lines.pop()
    return "\n".join(out_lines)

def _beautify_fields_for_excel(fields: Dict) -> Dict:
    # Copia defensiva
    out = dict(fields or {})

    # 1) Campos lista (SIEMPRE con guiones)
    bulleted_keys = [
        "criterios_adjudicacion_megadetallados",
        "sobre1_contenido",
        "sobre2_contenido",
        "sobre3_contenido",
    ]
    for k in bulleted_keys:
        v = str(out.get(k, "") or "")
        if v.strip():
            out[k] = _as_bulleted_multiline(v)

    # 2) Medios: SOLO listamos con guiones las líneas que ya son viñetas (perfiles).
    #    El texto explicativo debe quedar sin guion.
    v_medios = str(out.get("medios_humanos_materiales_detallados", "") or "")
    if v_medios.strip():
        out["medios_humanos_materiales_detallados"] = _as_mixed_multiline(v_medios)

    return out



def _set_cell(ws, cell: str, value: str) -> None:
    ws[cell].value = value
    try:
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
    except Exception:
        pass


def fill_template_excel(template_path: str, out_path: str, fields: Dict) -> None:
    wb = load_workbook(template_path)

    if "RESUMEN" not in wb.sheetnames:
        raise RuntimeError("La plantilla no tiene hoja 'RESUMEN'.")
    ws = wb["RESUMEN"]

    _set_cell(ws, "C2", fields.get("entidad_adjudicadora", ""))
    _set_cell(ws, "F2", fields.get("importe_licitacion", ""))
    _set_cell(ws, "I2", fields.get("fecha_limite", ""))
    _set_cell(ws, "F3", fields.get("plazo_ejecucion", ""))
    _set_cell(ws, "I3", fields.get("hora_limite", ""))

    _set_cell(ws, "C5", fields.get("titulo", ""))
    _set_cell(ws, "C6", fields.get("expediente", ""))
    _set_cell(ws, "C7", fields.get("cpv", ""))
    _set_cell(ws, "C8", fields.get("visita_tecnica", ""))
    _set_cell(ws, "C9", fields.get("lugar_prestacion", ""))

    _set_cell(ws, "B12", fields.get("objeto_detallado", ""))
    _set_cell(ws, "B15", fields.get("solvencia_tecnica_detallada", ""))
    _set_cell(ws, "B18", fields.get("medios_humanos_materiales_detallados", ""))
    _set_cell(ws, "B21", fields.get("criterios_adjudicacion_megadetallados", ""))

    if "SOBRES" in wb.sheetnames:
        ws2 = wb["SOBRES"]
        _set_cell(ws2, "C2", fields.get("sobre1_titulo", ""))
        _set_cell(ws2, "N2", fields.get("sobre1_pagina", ""))
        _set_cell(ws2, "C4", fields.get("sobre1_contenido", ""))

        _set_cell(ws2, "C8", fields.get("sobre2_titulo", ""))
        _set_cell(ws2, "N8", fields.get("sobre2_pagina", ""))
        _set_cell(ws2, "C10", fields.get("sobre2_contenido", ""))

        _set_cell(ws2, "C14", fields.get("sobre3_titulo", ""))
        _set_cell(ws2, "N14", fields.get("sobre3_pagina", ""))
        _set_cell(ws2, "C16", fields.get("sobre3_contenido", ""))

    wb.save(out_path)


# =========================================================
# MAIN: Generar Excel resumen IA usando una plantilla fija
# =========================================================
def generate_ai_summary_excel(
    tender_title: str,
    tender_link: str,
    template_folder: str,
    cache_folder: str,
    out_folder: str,
    manual_pcap_path: str,
    manual_ppt_path: Optional[str] = None,
    progress_cb: Optional[Callable[[float, str], None]] = None
) -> Tuple[str, str]:
    template_path = os.path.join(template_folder, "plantilla_resumen.xlsx")
    if not os.path.exists(template_path):
        raise RuntimeError(
            f"No encuentro la plantilla Excel en {template_path}. "
            "Pon tu Excel plantilla dentro de templates_resumen/ con nombre plantilla_resumen.xlsx"
        )

    if not manual_pcap_path or not os.path.exists(manual_pcap_path):
        raise RuntimeError("Falta PCAP (PDF) o no existe en disco.")

    ppt_exists = bool(manual_ppt_path and os.path.exists(manual_ppt_path))
    if progress_cb:
        try:
            progress_cb(0.10, "Leyendo PCAP…")
        except Exception:
            pass

    pcap_text = pdf_to_text_keep_pages(manual_pcap_path)
    if progress_cb:
        try:
            progress_cb(0.22, "PCAP leído. Preparando PPT…")
        except Exception:
            pass
    if len(re.sub(r"\s+", "", pcap_text)) < 800:
        raise RuntimeError("El PCAP parece no tener texto (posible PDF escaneado). Se requiere OCR.")

    ppt_text = ""
    if ppt_exists:
        ppt_text = pdf_to_text_keep_pages(manual_ppt_path)  # type: ignore
        if progress_cb:
            try:
                progress_cb(0.30, "PPT leído. Lanzando IA…")
            except Exception:
                pass
        if len(re.sub(r"\s+", "", ppt_text)) < 800:
            ppt_text = ""
            ppt_exists = False

    fields = llm_generate_excel_fields(
        tender_title=tender_title,
        tender_link=tender_link,
        pcap_text=pcap_text,
        ppt_text=ppt_text,
        progress_cb=progress_cb,
    )

    os.makedirs(out_folder, exist_ok=True)
    safe = re.sub(r"[^a-zA-Z0-9]+", "_", tender_title)[:60]
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(out_folder, f"RESUMENIA_{safe}_{ts}.xlsx")

    if progress_cb:
        try:
            progress_cb(0.95, "Escribiendo Excel…")
        except Exception:
            pass

    fields = _beautify_fields_for_excel(fields)

    fill_template_excel(template_path, out_path, fields)

    if progress_cb:
        try:
            progress_cb(1.0, "Excel generado")
        except Exception:
            pass

    info = "Excel Resumen IA generado usando plantilla fija"
    info += " | Con PPT" if ppt_exists else " | Sin PPT"
    info += " | GPT-5.4 (Responses): max_output_tokens=6500 (sin response_format por SDK antiguo)"
    info += " | Con control anti-referencias + corrección + vacíos sospechosos"
    return out_path, info